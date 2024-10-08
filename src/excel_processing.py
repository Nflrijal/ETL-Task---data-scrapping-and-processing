import os
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from database import create_table, execute_query
from logger_setup import setup_logger

# Setup the logger
logger = setup_logger(__name__)


def combine_tables_to_excel(response_folder, excel_file_path):
    """Function to combine the scrapped table data from the individual .html files
    into a single sheet in an excel file.

    :param response_folder: the folder to which the individual files are to be saved
    :type response_folder: str
    :param excel_file_path: name of the excel sheet
    :type excel_file_path: str

    """
    wb = Workbook()
    ws = wb.active
    ws.title = "NHL Stats 1990-2011"

    headers_written = False

    # Function to sort the file names based on the numeric values in the name
    def numerical_sort(value):
        return int(os.path.splitext(value)[0])

    sorted_filenames = sorted(
        os.listdir(response_folder), key=lambda x: numerical_sort(x)
    )

    for file_name in sorted_filenames:

        if file_name.endswith(".html"):
            file_path = os.path.join(response_folder, file_name)

            try:
                with open(file_path, "r", encoding="utf-8") as file:
                    soup = BeautifulSoup(file, "html.parser")
                    table = soup.find("table")

                    if not table:
                        logger.warning(f"No table found in file - {file_name}")
                        continue

                    rows = table.find_all("tr")
                    for row in rows:
                        cells = row.find_all(["td", "th"])
                        cell_values = [cell.get_text(strip=True) for cell in cells]

                        if not headers_written:
                            ws.append(cell_values)
                            headers_written = True
                        else:
                            if row.find("th"):
                                continue
                            ws.append(cell_values)
            except Exception as e:
                logger.error(f"Error processing file - {file_name}")
    try:
        wb.save(excel_file_path)
        logger.info(
            f"The table is saved to the excel file {excel_file_path} successfully..."
        )
    except Exception as e:
        logger.error(f"Failed to save table to excel file {excel_file_path}")


def write_summary_to_excel(excel_file_path, result, column_names):
    """Function to save the transformation result in to the excel file as new sheet

    :param excel_file_path: name of the excel sheet
    :type excel_file_path: str
    :param result: the transformation query resut to be saved
    :type result: str
    :param column_names: the column names from the query result
    :type result: list

    """

    try:

        wb = load_workbook(excel_file_path)

        summary_ws = wb.create_sheet(title="Winner and Loser per Year")

        summary_ws.append(column_names)

        for row in result:
            summary_ws.append(row)

        # Save the workbook
        wb.save(excel_file_path)
        logger.info(f"Summary written to {excel_file_path} successfully.")
    except Exception as e:
        logger.error(f"Error writing summary to Excel: {str(e)}")


def team_summary(excel_file_path, column_mapping, query):
    """Function to create the temporary table for performing the transformations to generate
    the teams summary.

    :param excel_file_path: the path of the source excel file
    :type excel_file_path: str
    :param column_mapping: the mapping for renaming the columns for creating the temp table
    :type column_mapping: dict
    :param query: the transformation query to be performed for generating the summary
    :type query: str

    """
    try:
        wb = load_workbook(excel_file_path)
        ws = wb.active
    except Exception as e:
        logger.error(f"Error loading the excel sheet {excel_file_path}")

    try:
        # creating the sqlite table
        columns = [column_mapping.get(cell.value, cell.value) for cell in ws[1]]
        create_table(ws, columns)
    except Exception as e:
        logger.error(f"Error creating the table with columns - {columns}  - {str(e)}")

    try:
        # performing the transformation
        result, query_columns = execute_query(query)
        if result and query_columns:
            write_summary_to_excel(excel_file_path, result, query_columns)
    except:
        logger.error(f"Error running query or writing summary to excel {str(e)}")
